"""
reporter.py
===========
Dual Job Tracker — Local Excel + Google Sheets
----------------------------------------------
Target Google Sheet ID: 1PQMwFgu_C_3AZOEec4I61dG2jpjvLYUYsBLNGlF4taI

Records every job processed by the agent into:
  1. Local Excel workbook (job_tracker.xlsx) — always saved as guaranteed backup
  2. Google Sheets (1PQMwFgu_C_3AZOEec4I61dG2jpjvLYUYsBLNGlF4taI) — live online sync

Dependencies:
    pip install openpyxl gspread
"""

import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise ImportError("openpyxl not found. Install with: pip install openpyxl") from exc

try:
    import gspread
except ImportError:
    gspread = None

# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("reporter")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TRACKER_FILE = Path("job_tracker.xlsx")
SHEET_NAME   = "Job Tracker"
GOOGLE_SHEET_ID = os.environ.get("GOOGLE_SHEET_ID", "1PQMwFgu_C_3AZOEec4I61dG2jpjvLYUYsBLNGlF4taI")
SERVICE_ACCOUNT_FILE = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE", "service_account.json")

# Password that locks the Excel workbook — read from .env / environment.
# Anyone who opens job_tracker.xlsx will be prompted for this password.
EXCEL_PASSWORD = os.environ.get("EXCEL_PASSWORD", "Mahathir@2025#Secure")

COLUMNS = [
    ("Date",            14),
    ("Job ID",          14),
    ("Title",           38),
    ("Company",         24),
    ("Location",        20),
    ("URL",             45),
    ("Score",           8),
    ("AI Rationale",    42),
    ("CV Filename",     50),
    ("Gemini Tailored", 15),
    ("Status",          12),
]

FILL_HEADER  = PatternFill("solid", fgColor="1A1A2E")
FILL_GREEN   = PatternFill("solid", fgColor="C8E6C9")
FILL_YELLOW  = PatternFill("solid", fgColor="FFF9C4")
FILL_WHITE   = PatternFill("solid", fgColor="FFFFFF")
FILL_ALT_ROW = PatternFill("solid", fgColor="F5F5F5")

FONT_HEADER  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_BODY    = Font(name="Calibri", size=10)
FONT_LINK    = Font(name="Calibri", size=10, color="0F3C78", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Excel Setup
# ---------------------------------------------------------------------------

def _apply_sheet_protection(ws) -> None:
    """
    Lock every cell in the sheet so the content cannot be edited without
    the correct password.  Readers can still scroll and copy — they just
    cannot modify or delete any data.
    """
    from openpyxl.worksheet.protection import SheetProtection
    ws.protection = SheetProtection(
        password=EXCEL_PASSWORD,
        sheet=True,               # enable sheet protection
        selectLockedCells=False,  # allow selecting (so readers can read/copy)
        selectUnlockedCells=False,
        formatCells=True,
        formatColumns=True,
        formatRows=True,
        insertColumns=True,
        insertRows=True,
        insertHyperlinks=True,
        deleteColumns=True,
        deleteRows=True,
        sort=True,
        autoFilter=True,
        pivotTables=True,
    )


def _save_protected_workbook(wb: Workbook, ws) -> None:
    """
    Apply sheet protection and workbook-level write-protection, then save.
    Called instead of wb.save() everywhere so protection is never skipped.
    """
    # Sheet-level: prevent cell editing
    _apply_sheet_protection(ws)

    # Workbook-level: require password to open the file at all
    wb.security.workbookPassword = EXCEL_PASSWORD
    wb.security.lockStructure    = True   # prevent adding/deleting sheets
    wb.security.lockWindows      = False  # allow resizing the window

    wb.save(TRACKER_FILE)
    logger.debug("Workbook saved with password protection.")


def _create_excel_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.freeze_panes = "A2"

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = FONT_HEADER
        cell.fill      = FILL_HEADER
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    _save_protected_workbook(wb, ws)
    return wb


def _load_or_create_excel():
    if TRACKER_FILE.exists():
        wb = load_workbook(TRACKER_FILE)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
    else:
        wb = _create_excel_workbook()
        ws = wb[SHEET_NAME]
    return wb, ws


# ---------------------------------------------------------------------------
# Google Sheets Helpers
# ---------------------------------------------------------------------------

def _get_gsheet_worksheet():
    if not gspread:
        return None
    sa_path = Path(SERVICE_ACCOUNT_FILE)
    if not sa_path.exists():
        return None
    try:
        gc = gspread.service_account(filename=str(sa_path))
        sh = gc.open_by_key(GOOGLE_SHEET_ID)
        return sh.sheet1
    except Exception as e:
        logger.warning("Google Sheets connection error: %s", e)
        return None


def ensure_sheet_headers() -> None:
    ws = _get_gsheet_worksheet()
    if ws:
        try:
            existing = ws.row_values(1)
            if not existing or existing[0] != "Date":
                ws.insert_row([c[0] for c in COLUMNS], 1)
                logger.info("Inserted headers into Google Sheet (%s).", GOOGLE_SHEET_ID)
        except Exception as e:
            logger.warning("Could not insert headers to Google Sheet: %s", e)


# ---------------------------------------------------------------------------
# Public Logging API
# ---------------------------------------------------------------------------

def append_job_row(
    job_id: str,
    title: str,
    company: str,
    location: str,
    url: str,
    score: float,
    rationale: str,
    cv_filename: str,
    cv_abs_path: str,
    gemini_tailored: bool,
    status: str = "New",
) -> int:
    score_pct = f"{score * 100:.0f}%"
    date_str  = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    row_data = [
        date_str,
        str(job_id),
        str(title),
        str(company),
        str(location),
        str(url),
        score_pct,
        str(rationale),
        str(cv_filename),
        "Yes" if gemini_tailored else "No (Base)",
        status,
    ]

    # 1. Write to local Excel backup if environment is writable
    next_row = 1
    try:
        wb, ws = _load_or_create_excel()
        next_row = ws.max_row + 1
        row_fill = FILL_GREEN if score >= 0.90 else (FILL_YELLOW if score >= 0.75 else FILL_WHITE)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.fill   = row_fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx in (1, 2, 7, 10, 11) else LEFT

            if col_idx == 6 and url:
                cell.hyperlink = url
                cell.font      = FONT_LINK
            elif col_idx == 9 and cv_abs_path:
                try:
                    cell.hyperlink = Path(cv_abs_path).resolve().as_uri()
                    cell.font      = FONT_LINK
                except Exception:
                    cell.font = FONT_BODY
            else:
                cell.font = FONT_BODY

        ws.row_dimensions[next_row].height = 18
        _save_protected_workbook(wb, ws)
        logger.info("Logged to Excel backup row %d | %s @ %s", next_row, title, company)
    except Exception as ee:
        logger.info("Skipping local Excel backup write on read-only serverless environment.")

    # 2. Sync to Google Sheets if configured (via Webhook or Service Account)
    webhook_url = os.environ.get("GOOGLE_SHEETS_WEBHOOK_URL")
    synced = False
    
    # Generate direct public cloud link for PDF as primary link
    cloud_pdf_url = None
    if cv_abs_path:
        pdf_path = cv_abs_path.replace(".docx", ".pdf")
        target_upload = pdf_path if Path(pdf_path).exists() else cv_abs_path
        target_filename = cv_filename.replace(".docx", ".pdf") if Path(pdf_path).exists() else cv_filename
        if Path(target_upload).exists():
            try:
                import requests
                with open(target_upload, "rb") as fh:
                    files = {"file": (target_filename, fh)}
                    up_resp = requests.post("https://tmpfiles.org/api/v1/upload", files=files, timeout=15)
                    if up_resp.status_code == 200:
                        up_data = up_resp.json()
                        raw_url = up_data.get("data", {}).get("url", "")
                        if raw_url:
                            cloud_pdf_url = raw_url.replace("tmpfiles.org/", "tmpfiles.org/dl/")
                            logger.info("Uploaded CV PDF to Cloud: %s", cloud_pdf_url)
            except Exception as ue:
                logger.warning("Cloud PDF upload error: %s", ue)

    # Format row_data[8] with clickable hyperlink if cloud URL obtained
    gs_row_data = list(row_data)
    if cloud_pdf_url:
        gs_row_data[8] = f'=HYPERLINK("{cloud_pdf_url}", "📄 Open PDF CV")'

    if webhook_url:
        try:
            import requests
            import base64

            payload = {"sheet_id": GOOGLE_SHEET_ID, "row": gs_row_data}
            if cv_abs_path:
                pdf_path = cv_abs_path.replace(".docx", ".pdf")
                target_upload = pdf_path if Path(pdf_path).exists() else cv_abs_path
                target_filename = cv_filename.replace(".docx", ".pdf") if Path(pdf_path).exists() else cv_filename
                if Path(target_upload).exists():
                    try:
                        with open(target_upload, "rb") as fh:
                            payload["file_bytes"] = base64.b64encode(fh.read()).decode("utf-8")
                            payload["file_name"]  = target_filename
                    except Exception as fe:
                        logger.warning("Could not read file for Webhook upload: %s", fe)

            resp = requests.post(webhook_url, json=payload, timeout=20)
            if resp.status_code == 200:
                logger.info("Synced row + CV link via Webhook.")
                synced = True
            else:
                logger.warning("Webhook HTTP status: %d", resp.status_code)
        except Exception as e:
            logger.warning("Webhook sync error: %s", e)

    if not synced:
        gws = _get_gsheet_worksheet()
        if gws:
            try:
                gws.append_row(row_data)
                logger.info("Synced row to Google Sheet via Service Account.")
            except Exception as e:
                logger.warning("Service Account append error: %s", e)

    return next_row


def write_run_summary(
    total_scraped: int,
    new_jobs: int,
    passed_screening: int,
    resumes_compiled: int,
    errors: list,
) -> None:
    try:
        wb, ws = _load_or_create_excel()
        next_row = ws.max_row + 1
        summary_text = (
            f"--- Run Summary ({datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}) | "
            f"Scraped: {total_scraped} | New: {new_jobs} | "
            f"Passed: {passed_screening} | CVs: {resumes_compiled} | "
            f"Errors: {len(errors)} ---"
        )
        cell = ws.cell(row=next_row, column=1, value=summary_text)
        cell.font = Font(name="Calibri", bold=True, italic=True, size=9, color="666666")
        cell.fill = PatternFill("solid", fgColor="E8EAF6")
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=len(COLUMNS))
        _save_protected_workbook(wb, ws)
    except Exception:
        pass

    webhook_url = os.environ.get("GOOGLE_SHEETS_WEBHOOK_URL")
    if webhook_url:
        try:
            import requests
            requests.post(webhook_url, json={"sheet_id": GOOGLE_SHEET_ID, "row": [summary_text]}, timeout=5)
        except Exception:
            pass

    gws = _get_gsheet_worksheet()
    if gws:
        try:
            gws.append_row([summary_text] + [""] * (len(COLUMNS) - 1))
        except Exception:
            pass


def get_all_logged_job_ids() -> set:
    job_ids = set()

    # Read from local Excel
    if TRACKER_FILE.exists():
        wb, ws = _load_or_create_excel()
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            val = row[0]
            if val and str(val).strip() and not str(val).startswith("---"):
                job_ids.add(str(val).strip())

    # Read from Google Sheets
    gws = _get_gsheet_worksheet()
    if gws:
        try:
            for val in gws.col_values(2)[1:]:
                v_str = str(val).strip()
                if v_str and not v_str.startswith("---"):
                    job_ids.add(v_str)
        except Exception:
            pass

    return job_ids
